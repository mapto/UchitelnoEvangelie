#!/usr/bin/env python3

from typing import List, Dict, Optional

from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import Font, colors  # type: ignore

from config import FROM_LANG, TO_LANG
from const import HILITE_PREFIX, IDX_COL, STYLE_COL, STYLE_SEP

from semantics import TableSemantics, MainLangSemantics, VarLangSemantics


def _style2str(s: Font, bgs: Dict[str, Optional[str]]) -> str:
    """take font style from index and presence of background in selected named columns"""
    result = [k for k, v in bgs.items() if v]
    if s and s.bold:
        result.append("bold")
    if s and s.italic:
        result.append("italic")
    return "|".join(result)


def import_mapping(fname: str, sem: TableSemantics) -> List[List[str]]:
    """Blank lines are perserved, as they serve as separators between groups (see merger).
    However, two subsequent blank lines are interpreted as end of file.
    """
    wb = load_workbook(fname, read_only=True, data_only=True)
    ws = wb.active
    assert ws, "non-null value required by mypy"

    result = []
    for row in ws.iter_rows(max_col=STYLE_COL):  # type: ignore
        line = [str(cell.value).strip() if cell.value else cell.value for cell in row]
        bgs = {
            f"{HILITE_PREFIX}{v:02d}{STYLE_SEP}{row[v].fill.start_color.rgb}": row[
                v
            ].fill.patternType
            for v in sem.cols()
            if row[v].fill and row[v].fill.start_color != colors.WHITE
        }
        line.append(_style2str(row[IDX_COL].font, bgs))
        result.append(line)

    return result


if __name__ == "__main__":
    sl_sem = MainLangSemantics(
        FROM_LANG, 4, [6, 7, 8, 9], VarLangSemantics(FROM_LANG, 0, [1, 2])
    )
    gr_sem = MainLangSemantics(
        TO_LANG, 10, [11, 12, 13], VarLangSemantics(TO_LANG, 15, [16, 17])
    )
    sem = TableSemantics(sl_sem, gr_sem)

    result = import_mapping("01-slovo1-lemat.xlsx", sem)

    print(result)
    print(len(result))
