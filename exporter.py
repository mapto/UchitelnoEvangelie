#!/usr/bin/env python3

from typing import List, Dict, Tuple

from openpyxl import Workbook  # type: ignore
from openpyxl.cell import WriteOnlyCell  # type: ignore
from openpyxl.styles import Font, PatternFill, Side, Border  # type: ignore
from openpyxl.styles.borders import BORDER_HAIR as BORDER_WIDTH  # type: ignore
from openpyxl.styles.fills import FILL_SOLID as FILL  # type: ignore

from model import Word

LIGHT_YELLOW = "FFFFE0"
DARK_GRAY = "A9A9A9"
LIGHT_GRAY = "D3D3D3"

BG_EDIT = PatternFill(start_color=LIGHT_YELLOW, fill_type=FILL)
BG_RESERVED = PatternFill(start_color=DARK_GRAY, fill_type=FILL)

_SIDE = Side(border_style=BORDER_WIDTH, color=LIGHT_GRAY)
BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


def row(row: Dict[str, str]) -> List[str]:
    result = [""] * 26
    for k, v in row.items():
        result[ord(k.upper()) - ord("A")] = v
    return result


def ifna_vlookup(ref: str, ret: str, rows: int) -> str:
    """
    https://support.microsoft.com/en-us/office/vlookup-function-0bbc8083-26fe-4963-8ab8-93a18ad188a1
    https://support.microsoft.com/en-us/office/ifna-function-6626c961-a569-42fc-a49d-79b4951fd461

    >>> ifna_vlookup("A3", "C", 5)
    '=_xlfn.IFNA(VLOOKUP(A3,A1:C5,3,FALSE),"")'
    """
    table_array = f"{ref[0]}1:{ret[0]}{rows}"
    col_index_num = ord(ret[0]) - ord(ref[0]) + 1
    return f'=_xlfn.IFNA(VLOOKUP({ref},{table_array},{col_index_num},FALSE),"")'


def if_isformula(ref: str) -> str:
    """
    https://support.microsoft.com/en-us/office/isformula-function-e4d1355f-7121-4ef2-801e-3839bfd6b1e5
    https://support.microsoft.com/en-us/office/if-function-69aed7c9-4e8a-4755-a9bc-aa8bbff73be2

    >>> if_isformula("B3")
    '=IF(_xlfn.ISFORMULA(B3),"",B3)'
    """
    return f'=IF(_xlfn.ISFORMULA({ref}),"",{ref})'


def export_sheet(data: List[Word], fname: str) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    total = len(data)
    for i, next in enumerate(data):
        r = i + 1
        vref = f"A{r}"
        bref = f"E{r}"
        gref = f"K{r}"
        content = {
            "A": next.variant,
            "B": ifna_vlookup(vref, "S", total),
            "D": next.index(),
            "E": next.word,
            "F": next.line_context,
            "G": ifna_vlookup(bref, "T", total),
            "H": ifna_vlookup(bref, "U", total),
            "I": ifna_vlookup(bref, "V", total),
            "J": ifna_vlookup(bref, "W", total),
            "L": ifna_vlookup(gref, "X", total),
            "M": ifna_vlookup(gref, "Y", total),
            "N": ifna_vlookup(gref, "Z", total),
            "S": if_isformula(f"B{r}"),
            "T": if_isformula(f"G{r}"),
            "U": if_isformula(f"H{r}"),
            "V": if_isformula(f"I{r}"),
            "W": if_isformula(f"J{r}"),
            "X": if_isformula(f"L{r}"),
            "Y": if_isformula(f"M{r}"),
            "Z": if_isformula(f"N{r}"),
        }

        line = []
        for v in row(content):
            cell = WriteOnlyCell(ws, value=v)
            if cell.value:
                cell.font = Font("CyrillicaOchrid10U")
                cell.border = BORDER
                if cell.value.startswith("=IF"):
                    cell.fill = BG_RESERVED
                elif cell.value.startswith("=_xlfn.IFNA"):
                    cell.fill = BG_EDIT
            line.append(cell)
        ws.append(line)

    # Save the file
    wb.save(fname)
