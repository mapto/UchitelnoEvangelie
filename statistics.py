#!/usr/bin/env python3
"""Table counter for Uchitelno Evangelie.
Licensed under MIT License, detailed here: https://mit-license.org/

Usage:
  integrator.py [-v|--verbose|-s|--silent] <xlsx>...

Options:
  -h --help             This information
  -s --silent           Remove output other than warnings and errors, also does not pause after completion
  -v --verbose          Increase debug level
  --version             Print version

"""
import csv

from docopt import docopt  # type: ignore
from openpyxl import load_workbook  # type: ignore

__version__ = "0.0.1"  # used also by build.sh script

if __name__ == "__main__":
    args = docopt(__doc__, version=__version__)
    """Blank lines are perserved, as they serve as separators between groups (see merger).
    However, two subsequent blank lines are interpreted as end of file.
    """
    fnames = sorted(args["<xlsx>"])
    counters = []
    for fname in fnames:
        print(fname)
        wb = load_workbook(fname, read_only=True, data_only=True)
        ws = wb.active

        columns = [
            "славянски вариант употреба",
            "славянски вариант лема",
            "славянски вариант подлема",
            "славянски вариант втора подлема",
            "адрес",
            "славянски основен употреба",
            "ред",
            "славянски основен лема",
            "славянски основен подлема",
            "славянски основен втора подлема",
            "славянски основен трета подлема",
            "гръцки основен употреба",
            "гръцки основен лема",
            "гръцки основен подлема",
            "гръцки основен втора подлема",
            "гръцки основен трета подлема",
            "гръцки вариант употреба",
            "гръцки вариант лема",
            "гръцки вариант подлема",
            "гръцки вариант втора подлема",
            "гръцки вариант трета подлема",
        ]
        counter = [0] * len(columns)
        blank = False
        for row in ws.iter_rows(max_col=len(columns)):
            line = [
                str(cell.value).strip() if cell.value else cell.value for cell in row
            ]
            # Two consequent blank lines
            if blank and not [l for l in line if l]:
                break

            for i, c in enumerate(line):
                if c:
                    counter[i] += 1

            blank = not [l for l in line if l]
        counters += [counter]

    with open("statistics.csv", "w") as fout:
        w = csv.writer(fout)
        w.writerow(["слово"] + columns)
        for i in range(len(fnames)):
            w.writerow([fnames[i].split("/")[-1]] + counters[i])
