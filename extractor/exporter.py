#!/usr/bin/env python3

from typing import List, Dict, Tuple

from openpyxl import Workbook  # type: ignore
from openpyxl.cell import WriteOnlyCell  # type: ignore
from openpyxl.styles import Font, PatternFill, Side, Border  # type: ignore
from openpyxl.styles.borders import BORDER_HAIR as BORDER_WIDTH  # type: ignore
from openpyxl.styles.fills import FILL_SOLID as FILL  # type: ignore

from model import Word, WordList

# LIGHT_YELLOW = "FFFFE0"
DARK_GRAY = "A9A9A9"
LIGHT_GRAY = "D3D3D3"

# BG_EDIT = PatternFill(start_color=LIGHT_YELLOW, fill_type=FILL)
BG_RESERVED = PatternFill(start_color=DARK_GRAY, fill_type=FILL)

_SIDE = Side(border_style=BORDER_WIDTH, color=LIGHT_GRAY)
BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
COLUMN_BOUND = 30
LETTERS = 26

def row(row: Dict[int, str]) -> List[str]:
    result = [""] * 35
    for k, v in row.items():
        result[k] = v
    return result


def sheet_column(i: int) -> str:
    """
    Converts a 0-index of column to the spreadsheet column name
    >>> sheet_column(0)
    'A'
    >>> sheet_column(25)
    'Z'
    >>> sheet_column(26)
    'AA'
    >>> sheet_column(51)
    'AZ'
    >>> sheet_column(52)
    'BA'
    """
    if i >= LETTERS:
        return chr(ord('A') - 1 + i // LETTERS) + chr(ord('A') + i % LETTERS)
    return chr(ord('A') + i)

# def ifna_vlookup_old(ref: str, ret: str, rows: int) -> str:
#     """
#     https://support.microsoft.com/en-us/office/vlookup-function-0bbc8083-26fe-4963-8ab8-93a18ad188a1
#     https://support.microsoft.com/en-us/office/ifna-function-6626c961-a569-42fc-a49d-79b4951fd461

#     >>> ifna_vlookup("A3", "C", 5)
#     '=_xlfn.IFNA(VLOOKUP(A3,A1:C5,3,FALSE),"")'
#     """
#     table_array = f"{ref[0]}1:{ret[0]}{rows}"
#     col_index_num = ord(ret[0]) - ord(ref[0]) + 1
#     return f'=_xlfn.IFNA(VLOOKUP({ref},{table_array},{col_index_num},FALSE),"")'


def ifna_vlookup(val_col: int, row: int, ret_col: int, rows: int) -> str:
    """
    https://support.microsoft.com/en-us/office/vlookup-function-0bbc8083-26fe-4963-8ab8-93a18ad188a1
    https://support.microsoft.com/en-us/office/ifna-function-6626c961-a569-42fc-a49d-79b4951fd461

    >>> ifna_vlookup(0, 3, 2, 5)
    '=_xlfn.IFNA(VLOOKUP(A3,A1:C5,3,FALSE),"")'

    >>> ifna_vlookup(5, 1, 26, 10)
    '=_xlfn.IFNA(VLOOKUP(F1,F1:AA10,22,FALSE),"")'
    """
    ref = f"{sheet_column(val_col)}{row}"
    table_array = f"{sheet_column(val_col)}1:{sheet_column(ret_col)}{rows}"
    col_index_num = ret_col - val_col + 1
    return f'=_xlfn.IFNA(VLOOKUP({ref},{table_array},{col_index_num},FALSE),"")'


def if_isformula(ref: str) -> str:
    """
    https://support.microsoft.com/en-us/office/isformula-function-e4d1355f-7121-4ef2-801e-3839bfd6b1e5
    https://support.microsoft.com/en-us/office/if-function-69aed7c9-4e8a-4755-a9bc-aa8bbff73be2

    >>> if_isformula("B3")
    '=IF(_xlfn.ISFORMULA(B3),"",B3)'
    """
    return f'=IF(_xlfn.ISFORMULA({ref}),"",{ref})'

def semantics(count: int, word: int, lem_start: int, aux_start: int, ref: int, rows: int)-> Dict[int, str]:
    """
    Slavic variants
    >>> semantics(3, 0, 1, 20, 1, 10)
    {1: '=_xlfn.IFNA(VLOOKUP(A1,A1:U10,21,FALSE),"")', 20: '=IF(_xlfn.ISFORMULA(B1),"",B1)', 2: '=_xlfn.IFNA(VLOOKUP(A1,A1:V10,22,FALSE),"")', 21: '=IF(_xlfn.ISFORMULA(C1),"",C1)', 3: '=_xlfn.IFNA(VLOOKUP(A1,A1:W10,23,FALSE),"")', 22: '=IF(_xlfn.ISFORMULA(D1),"",D1)'}

    Main slavic    
    >>> semantics(4, 5, 7, 23, 1, 10)
    {7: '=_xlfn.IFNA(VLOOKUP(F1,F1:X10,19,FALSE),"")', 23: '=IF(_xlfn.ISFORMULA(H1),"",H1)', 8: '=_xlfn.IFNA(VLOOKUP(F1,F1:Y10,20,FALSE),"")', 24: '=IF(_xlfn.ISFORMULA(I1),"",I1)', 9: '=_xlfn.IFNA(VLOOKUP(F1,F1:Z10,21,FALSE),"")', 25: '=IF(_xlfn.ISFORMULA(J1),"",J1)', 10: '=_xlfn.IFNA(VLOOKUP(F1,F1:AA10,22,FALSE),"")', 26: '=IF(_xlfn.ISFORMULA(K1),"",K1)'}
    
    Variant greek
    >>> semantics(4, 15, 16, 29, 1, 10)
    {16: '=_xlfn.IFNA(VLOOKUP(P1,P1:AD10,15,FALSE),"")', 29: '=IF(_xlfn.ISFORMULA(Q1),"",Q1)', 17: '=_xlfn.IFNA(VLOOKUP(P1,P1:AE10,16,FALSE),"")', 30: '=IF(_xlfn.ISFORMULA(R1),"",R1)', 18: '=_xlfn.IFNA(VLOOKUP(P1,P1:AF10,17,FALSE),"")', 31: '=IF(_xlfn.ISFORMULA(S1),"",S1)', 19: '=_xlfn.IFNA(VLOOKUP(P1,P1:AG10,18,FALSE),"")', 32: '=IF(_xlfn.ISFORMULA(T1),"",T1)'}
    """
    result = {}
    for i in range(count):
        # ref = word + i
        lem = lem_start + i
        aux = aux_start + i
        # result[lem] = ifna_vlookup(f"{sheet_column(word)}{ref}", sheet_column(aux), rows)
        result[lem] = ifna_vlookup(word, ref, aux, rows)
        result[aux] = if_isformula(f"{sheet_column(lem)}{ref}")
    return result


def export_sheet(data: WordList, fname: str) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    total = len(data)
    for i, next in enumerate(data):
        content = {
            0: next.variant,
            4: next.index(),
            5: next.word,
            6: next.line_context,
        }
        # slavic variants
        content.update(semantics(3, 0, 1, 20, i + 1, total))
        # slavic main
        content.update(semantics(4, 5, 7, 23, i + 1, total))
        # greek main
        content.update(semantics(4, 11, 12, 27, i + 1, total))
        # greek variants
        content.update(semantics(3, 16, 17, 31, i + 1, total))

        line = []
        for v in row(content):
            cell = WriteOnlyCell(ws, value=v)
            if cell.value:
                cell.font = Font("CyrillicaOchrid10U")
                cell.border = BORDER
                if cell.value.startswith("=IF"):
                    cell.fill = BG_RESERVED
                # elif cell.value.startswith("=_xlfn.IFNA"):
                    # cell.fill = BG_EDIT
            line.append(cell)
        ws.append(line)

    # Save the file
    wb.save(fname)
